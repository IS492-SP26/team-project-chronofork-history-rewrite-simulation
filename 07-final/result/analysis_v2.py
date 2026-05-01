"""
Pilot Study Quantitative Analysis v2
- Adds Wilcoxon signed-rank tests
- Reorganizes visualizations by Pilot Questionnaire categories (A1, A2, B1, B2, D)
- Charts are smaller / clustered to serve presentation narrative
"""

import os
import json
import numpy as np
import openpyxl
from scipy import stats
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

# ── Colors (UIUC palette) ──
ORANGE = '#E84A27'
BLUE = '#13294B'
LIGHT_ORANGE = '#F4A582'
LIGHT_BLUE = '#92C5DE'
GRAY = '#7A7A7A'

# ── Paths ──
BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(os.path.dirname(BASE), 'raw_data', 'questionaire.xlsx')
CHART_DIR = os.path.join(BASE, 'charts')
os.makedirs(CHART_DIR, exist_ok=True)

# ── Load data ──
wb = openpyxl.load_workbook(XLSX)
ws = wb.active

rows = []
for r in range(2, ws.max_row + 1):
    system = ws.cell(row=r, column=1).value
    items = [ws.cell(row=r, column=c).value for c in range(2, 50)]
    rows.append((system, items))

participants = []
for i in range(0, len(rows), 2):
    r1_sys, r1_items = rows[i]
    r2_sys, r2_items = rows[i + 1]
    pid = f"P{i // 2 + 1}"
    if r1_sys == 'Chrono':
        participants.append({'id': pid, 'Chrono': r1_items, 'IDN': r2_items})
    else:
        participants.append({'id': pid, 'Chrono': r2_items, 'IDN': r1_items})

# ── Reverse scoring ──
REVERSE_ITEMS_7PT = [1, 2, 3, 4, 5, 6, 25, 32, 33, 44, 46]

def reverse_score(items):
    scored = list(items)
    for idx in REVERSE_ITEMS_7PT:
        i = idx - 1
        scored[i] = 8 - scored[i]
    return scored

for p in participants:
    p['Chrono_scored'] = reverse_score(p['Chrono'])
    p['IDN_scored'] = reverse_score(p['IDN'])

# ── Subscales ──
subscales = {
    'Narrative Understanding': {'items': [1, 2, 3], 'cat': 'A1'},
    'Attentional Focus':       {'items': [4, 5, 6], 'cat': 'A1'},
    'Narrative Presence':      {'items': [7, 8, 9], 'cat': 'A1'},
    'Emotional Engagement':    {'items': [10, 11, 12], 'cat': 'A1'},
    'NES Overall':             {'items': list(range(1, 13)), 'cat': 'A1'},
    'Explanation Quality':     {'items': list(range(13, 20)), 'cat': 'A2'},
    'Alt. Path Plausibility':  {'items': [20], 'cat': 'A2'},
    'Agency (PENS Core)':      {'items': [21, 22, 23], 'cat': 'B1'},
    'PENS Extended':           {'items': [21, 22, 23, 24, 25], 'cat': 'B1'},
    'UMUX-Lite':               {'items': [26, 27], 'cat': 'B2'},
    'Mental Effort (Paas)':    {'items': [28], 'cat': 'B2', 'scale': 9},
    'Perspective Switching':   {'items': [29], 'cat': 'B2'},
    'Interest/Enjoyment':      {'items': list(range(30, 37)), 'cat': 'D'},
    'Value/Usefulness':        {'items': list(range(37, 44)), 'cat': 'D'},
    'Pressure/Tension':        {'items': list(range(44, 49)), 'cat': 'D'},
}

def compute(scored, idxs):
    return float(np.mean([scored[i - 1] for i in idxs]))

results = {}
for name, spec in subscales.items():
    chrono = [compute(p['Chrono_scored'], spec['items']) for p in participants]
    idn = [compute(p['IDN_scored'], spec['items']) for p in participants]
    results[name] = {'Chrono': chrono, 'IDN': idn}

# ── Wilcoxon signed-rank tests ──
print("=" * 88)
print("WILCOXON SIGNED-RANK TESTS (paired, two-sided)")
print("=" * 88)
print(f"{'Subscale':30s} | {'Mdiff':>7s} | {'W':>6s} | {'p':>7s} | {'r':>6s} | direction")
print("-" * 88)

test_results = {}
for name in subscales:
    c = np.array(results[name]['Chrono'])
    i = np.array(results[name]['IDN'])
    diff = c - i
    n_nonzero = np.sum(diff != 0)
    if n_nonzero < 2:
        test_results[name] = {'W': None, 'p': None, 'r': None,
                              'mean_diff': float(np.mean(diff))}
        continue
    try:
        # Use pratt method to handle ties (or wilcox to drop zeros)
        res = stats.wilcoxon(c, i, zero_method='wilcox', alternative='two-sided',
                             correction=False, method='exact' if n_nonzero <= 10 else 'approx')
        W = float(res.statistic)
        p = float(res.pvalue)
    except Exception as e:
        res = stats.wilcoxon(c, i, zero_method='wilcox', alternative='two-sided')
        W = float(res.statistic)
        p = float(res.pvalue)
    # Effect size r = Z / sqrt(N)  — approx using normal approx
    try:
        z_res = stats.wilcoxon(c, i, zero_method='wilcox', alternative='two-sided', method='approx')
        # method='approx' returns p-value from normal approximation; back out Z from p
        from scipy.stats import norm
        z = norm.isf(z_res.pvalue / 2) * (1 if np.mean(diff) > 0 else -1)
        r = z / np.sqrt(n_nonzero)
    except Exception:
        r = None
    mean_diff = float(np.mean(diff))
    direction = "Chrono >" if mean_diff > 0 else ("IDN >" if mean_diff < 0 else "tie")
    test_results[name] = {'W': W, 'p': p, 'r': r, 'mean_diff': mean_diff, 'n_nonzero': int(n_nonzero)}
    r_str = f"{r:+.2f}" if r is not None else "  n/a"
    print(f"{name:30s} | {mean_diff:+7.2f} | {W:6.1f} | {p:7.4f} | {r_str:>6s} | {direction}")

# Save test results
with open(os.path.join(BASE, 'wilcoxon_results.json'), 'w') as f:
    json.dump(test_results, f, indent=2)

# ────────────────────────────────────────────────────────────────────────
#  PRESENTATION-ALIGNED CHARTS
#  Each chart serves one slide-narrative beat.
# ────────────────────────────────────────────────────────────────────────

plt.rcParams.update({
    'font.family': 'DejaVu Sans',
    'font.size': 11,
    'axes.titlesize': 13,
    'axes.titleweight': 'bold',
    'axes.labelsize': 11,
    'figure.facecolor': 'white',
    'axes.facecolor': 'white',
    'savefig.dpi': 220,
    'savefig.bbox': 'tight',
})

def desc(vs):
    return float(np.mean(vs)), float(np.std(vs, ddof=1))

def annotate_sig(ax, x_pos, y_top, p):
    """Place a significance asterisk if p<.05, dot if p<.10."""
    if p is None:
        return
    if p < 0.001:
        sym = '***'
    elif p < 0.01:
        sym = '**'
    elif p < 0.05:
        sym = '*'
    elif p < 0.10:
        sym = '†'
    else:
        return
    ax.text(x_pos, y_top, sym, ha='center', va='bottom', fontsize=14, fontweight='bold')

def paired_bar_panel(ax, names, *, title, ylim=(0, 7.5), ylabel='Mean (1–7)', show_p=True):
    """Bar chart of Chrono vs IDN means for the given subscales."""
    x = np.arange(len(names))
    w = 0.36
    c_means = [desc(results[n]['Chrono'])[0] for n in names]
    c_sds = [desc(results[n]['Chrono'])[1] for n in names]
    i_means = [desc(results[n]['IDN'])[0] for n in names]
    i_sds = [desc(results[n]['IDN'])[1] for n in names]

    ax.bar(x - w/2, c_means, w, yerr=c_sds, label='ChronoFork', color=ORANGE,
           edgecolor='white', linewidth=0.8, capsize=3, error_kw={'linewidth': 0.9, 'ecolor': '#444'})
    ax.bar(x + w/2, i_means, w, yerr=i_sds, label='IDN-Twine', color=BLUE,
           edgecolor='white', linewidth=0.8, capsize=3, error_kw={'linewidth': 0.9, 'ecolor': '#444'})

    # Significance markers above the higher bar
    if show_p:
        for j, n in enumerate(names):
            top = max(c_means[j] + c_sds[j], i_means[j] + i_sds[j]) + 0.25
            annotate_sig(ax, j, top, test_results[n]['p'])

    ax.set_xticks(x)
    ax.set_xticklabels([n.replace(' (PENS Core)', '').replace('Engagement', 'Engage')
                        .replace('Understanding', 'Underst.') for n in names], fontsize=9.5)
    ax.set_ylim(*ylim)
    ax.set_ylabel(ylabel)
    ax.set_title(title, loc='left')
    ax.axhline(y=4, color=GRAY, linestyle='--', alpha=0.35, linewidth=0.8)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.legend(frameon=False, loc='upper right', fontsize=9)

# ── Chart 1: A1 Narrative Engagement (4 subscales + overall) ──
fig, ax = plt.subplots(figsize=(7.5, 4.2))
A1_subs = ['Narrative Understanding', 'Attentional Focus', 'Narrative Presence', 'Emotional Engagement', 'NES Overall']
paired_bar_panel(ax, A1_subs, title='A1  Narrative Engagement (NES, 1–7)')
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'A1_narrative_engagement.png'))
plt.close(fig)
print("\nSaved: A1_narrative_engagement.png")

# ── Chart 2: A2 Explanation Quality + Plausibility ──
fig, ax = plt.subplots(figsize=(5.5, 4.2))
A2_subs = ['Explanation Quality', 'Alt. Path Plausibility']
paired_bar_panel(ax, A2_subs, title='A2  Explanation & Plausibility (1–7)')
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'A2_explanation_plausibility.png'))
plt.close(fig)
print("Saved: A2_explanation_plausibility.png")

# ── Chart 3: B1 Agency ──
fig, ax = plt.subplots(figsize=(5.5, 4.2))
B1_subs = ['Agency (PENS Core)', 'PENS Extended']
paired_bar_panel(ax, B1_subs, title='B1  Sense of Agency (PENS, 1–7)')
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'B1_agency.png'))
plt.close(fig)
print("Saved: B1_agency.png")

# ── Chart 4: B2 Usability cluster (UMUX-Lite + Perspective Switching) — NOT mental effort ──
fig, ax = plt.subplots(figsize=(5.5, 4.2))
B2_likert = ['UMUX-Lite', 'Perspective Switching']
paired_bar_panel(ax, B2_likert, title='B2  Usability (1–7)')
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'B2_usability.png'))
plt.close(fig)
print("Saved: B2_usability.png")

# ── Chart 5: B2 Mental Effort (separate, 1–9 scale) ──
fig, ax = plt.subplots(figsize=(5.0, 4.2))
ME = test_results['Mental Effort (Paas)']
c_me = results['Mental Effort (Paas)']['Chrono']
i_me = results['Mental Effort (Paas)']['IDN']
c_m, c_s = desc(c_me)
i_m, i_s = desc(i_me)
xs = [0, 1]
bars = ax.bar(xs, [c_m, i_m], yerr=[c_s, i_s], width=0.55,
              color=[ORANGE, BLUE], edgecolor='white', linewidth=0.8,
              capsize=4, error_kw={'linewidth': 1.0, 'ecolor': '#444'})
# Per-participant scatter
for j, p in enumerate(participants):
    ax.scatter([0 + np.random.uniform(-0.08, 0.08)], [c_me[j]], color='white',
               edgecolor='black', s=30, zorder=3, linewidth=0.6)
    ax.scatter([1 + np.random.uniform(-0.08, 0.08)], [i_me[j]], color='white',
               edgecolor='black', s=30, zorder=3, linewidth=0.6)
ax.set_xticks(xs)
ax.set_xticklabels(['ChronoFork', 'IDN-Twine'])
ax.set_ylim(0, 10)
ax.set_ylabel('Paas Mental Effort  (1 = very low, 9 = very high)')
ax.set_title('B2  Cognitive Load (Paas, 1–9)', loc='left')
ax.axhline(y=5, color=GRAY, linestyle='--', alpha=0.35, linewidth=0.8)
annotate_sig(ax, 0.5, max(c_m + c_s, i_m + i_s) + 0.7, ME['p'])
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
# Add ΔM annotation
ax.text(0.5, 9.5, f"ΔM = +{ME['mean_diff']:.2f}   p = {ME['p']:.3f}",
        ha='center', va='top', fontsize=9.5, color='#333',
        bbox=dict(boxstyle='round,pad=0.3', facecolor='#FFF8F2', edgecolor='#E8C9B7'))
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'B2_mental_effort.png'))
plt.close(fig)
print("Saved: B2_mental_effort.png")

# ── Chart 6: D Motivation (IMI: Interest, Value, Pressure) ──
fig, ax = plt.subplots(figsize=(6.0, 4.2))
D_subs = ['Interest/Enjoyment', 'Value/Usefulness', 'Pressure/Tension']
paired_bar_panel(ax, D_subs, title='D  Motivation (IMI, 1–7)')
# Annotate that lower=better for Pressure
ax.text(2.0, 0.4, '↓ lower = less pressure', ha='center', fontsize=8, color='#666', style='italic')
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'D_motivation.png'))
plt.close(fig)
print("Saved: D_motivation.png")

# ── Chart 7: Headline summary radar (5 narrative-aligned dimensions) ──
radar_subs = ['NES Overall', 'Explanation Quality', 'Agency (PENS Core)',
              'Interest/Enjoyment', 'Value/Usefulness']
N = len(radar_subs)
angles = np.linspace(0, 2 * np.pi, N, endpoint=False).tolist()
angles += angles[:1]
c_vals = [desc(results[n]['Chrono'])[0] for n in radar_subs]; c_vals += c_vals[:1]
i_vals = [desc(results[n]['IDN'])[0] for n in radar_subs]; i_vals += i_vals[:1]

fig, ax = plt.subplots(figsize=(6.5, 6.5), subplot_kw=dict(polar=True))
ax.set_theta_offset(np.pi / 2)
ax.set_theta_direction(-1)
ax.plot(angles, c_vals, 'o-', color=ORANGE, linewidth=2, label='ChronoFork', markersize=7)
ax.fill(angles, c_vals, alpha=0.18, color=ORANGE)
ax.plot(angles, i_vals, 's-', color=BLUE, linewidth=2, label='IDN-Twine', markersize=7)
ax.fill(angles, i_vals, alpha=0.15, color=BLUE)
ax.set_xticks(angles[:-1])
ax.set_xticklabels([n.replace(' (PENS Core)', '') for n in radar_subs], fontsize=10)
ax.set_ylim(0, 7)
ax.set_yticks([2, 4, 6])
ax.set_yticklabels(['2', '4', '6'], fontsize=8)
ax.set_title('Headline Profile  (5 dimensions, 1–7)', y=1.10, fontsize=13, fontweight='bold')
ax.legend(loc='upper right', bbox_to_anchor=(1.25, 1.10), frameon=False)
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'headline_radar.png'))
plt.close(fig)
print("Saved: headline_radar.png")

# ── Chart 8: Per-participant slope chart for THE single most-important dimension (Agency) ──
# Plus a small one for Mental Effort to support the "germane load" story.
def slope(ax, name, *, title, scale=7, lower_is_better=False):
    c = results[name]['Chrono']
    i = results[name]['IDN']
    for j, p in enumerate(participants):
        ax.plot([0, 1], [c[j], i[j]], 'o-', color=GRAY, alpha=0.55, markersize=5, linewidth=1)
        ax.text(-0.05, c[j], p['id'], ha='right', va='center', fontsize=8, color='#555')
    ax.scatter([0]*len(c), c, color=ORANGE, s=50, zorder=3, edgecolor='white', linewidth=0.8)
    ax.scatter([1]*len(i), i, color=BLUE, s=50, zorder=3, edgecolor='white', linewidth=0.8)
    ax.plot([0, 1], [np.mean(c), np.mean(i)], color='black', linewidth=2.2, marker='D',
            markerfacecolor='white', markeredgewidth=1.5, markersize=9, zorder=4)
    ax.set_xticks([0, 1])
    ax.set_xticklabels(['ChronoFork', 'IDN-Twine'])
    ax.set_xlim(-0.25, 1.2)
    ax.set_ylim(0, scale + 0.5)
    ax.set_ylabel(f'1–{scale}')
    tr = test_results[name]
    p_str = f"p = {tr['p']:.3f}" if tr['p'] is not None else "p = n/a"
    ax.set_title(f"{title}\nΔM = {tr['mean_diff']:+.2f}   {p_str}", loc='left', fontsize=11)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

fig, axes = plt.subplots(1, 2, figsize=(9, 4.5))
slope(axes[0], 'Agency (PENS Core)', title='Agency  (PENS Core)')
slope(axes[1], 'Mental Effort (Paas)', title='Mental Effort  (Paas, 1–9)', scale=9)
fig.tight_layout()
fig.savefig(os.path.join(CHART_DIR, 'slope_agency_effort.png'))
plt.close(fig)
print("Saved: slope_agency_effort.png")

# ── Per-participant table (markdown) ──
with open(os.path.join(BASE, 'per_participant_v2.md'), 'w') as f:
    f.write("# Per-Participant Subscale Scores (with Wilcoxon)\n\n")
    f.write("| Subscale | " + " | ".join(p['id'] for p in participants) +
            " | Mean Δ | W | p | r |\n")
    f.write("|" + "---|" * (len(participants) + 5) + "\n")
    for n in subscales:
        diffs = [c - i for c, i in zip(results[n]['Chrono'], results[n]['IDN'])]
        tr = test_results[n]
        line = f"| {n} | " + " | ".join(f"{d:+.2f}" for d in diffs) + \
               f" | {tr['mean_diff']:+.2f} | {tr['W']:.1f} | {tr['p']:.3f} | "
        line += (f"{tr['r']:+.2f}" if tr['r'] is not None else "n/a") + " |\n"
        f.write(line)
print("Saved: per_participant_v2.md")
print("\nDone.")
